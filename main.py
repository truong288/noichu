from telegram import Update  #ok chạy đa nhóm có gắn link CARO
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
import asyncio
import re
import json
import os
from datetime import datetime
import openpyxl
from stay_alive import keep_alive

keep_alive()

players = {}
player_names = {}
player_usernames = {}
player_join_times = {}
current_phrase = {}
used_phrases = {}
current_player_index = {}
in_game = {}
waiting_for_phrase = {}
turn_timeout_task = {}
game_start_time = {}
chat_id = {}
BANNED_USERS = {}


GLOBAL_BANNED_WORDS = {
    "đần", "bần", "ngu", "ngốc", "bò", "dốt", "nát", "chó", "địt", "mẹ", "mày",
    "chi", "mô", "răng", "rứa", "má", "lồn", "lòn", "cứt"
}
BANNED_WORDS_FILE = "banned_words.txt"

def load_banned_words():
    if os.path.exists(BANNED_WORDS_FILE):
        with open(BANNED_WORDS_FILE, "r", encoding="utf-8") as f:
            return set(line.strip().lower() for line in f.readlines())
    return set()

def save_banned_words():
    with open(BANNED_WORDS_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(sorted(GLOBAL_BANNED_WORDS)))

GLOBAL_BANNED_WORDS.update(load_banned_words())

STATS_FILE = "winners.json"
EXCEL_FILE = "danh_sach.xlsx"


def load_tu_don():
    if not os.path.exists("tu_don.txt"):
        return set()
    with open("tu_don.txt", "r", encoding="utf-8") as f:
        return set(word.strip().lower() for word in f.readlines()
                   if word.strip())


TU_DON_LIST = load_tu_don()


def is_admin(user_id):
    admin_ids = [5429428390, 5930936939, 7034158998]
    return user_id in admin_ids


def load_stats():
    if os.path.exists(STATS_FILE):
        with open(STATS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_stats(data):
    with open(STATS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


stats = load_stats()


def save_player_to_excel(user_id, name, username, join_time):
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if today in wb.sheetnames:
            ws = wb[today]
        else:
            ws = wb.create_sheet(today)
            ws.append([
                "Tên người chơi", "Username", "Telegram ID",
                "Thời gian tham gia"
            ])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = today
        ws.append([
            "Tên người chơi", "Username", "Telegram ID", "Thời gian tham gia"
        ])

    if user_id not in [row[2].value for row in ws.iter_rows(min_row=2)]:
        ws.append([name, username, user_id, join_time])
        wb.save(EXCEL_FILE)


def reset_game_state(chat_id):
    global players, player_names, player_usernames, player_join_times, current_phrase, used_phrases, current_player_index, in_game, waiting_for_phrase, turn_timeout_task, game_start_time
    players[chat_id] = []
    player_names[chat_id] = {}
    player_usernames[chat_id] = {}
    player_join_times[chat_id] = {}
    current_phrase[chat_id] = ""
    used_phrases[chat_id] = {}
    current_player_index[chat_id] = 0
    in_game[chat_id] = False
    waiting_for_phrase[chat_id] = False
    game_start_time[chat_id] = None
    if turn_timeout_task.get(chat_id):
        turn_timeout_task[chat_id].cancel()
        turn_timeout_task[chat_id] = None


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user

    # Nếu là ADMIN → Reset toàn bộ
    if is_admin(user.id):
        global stats
        stats = {}  # Xóa toàn bộ stats
        save_stats(stats)
        
        # Reset trạng thái tất cả nhóm
        for group_id in list(players.keys()):
            reset_game_state(group_id)
        
        await update.message.reply_text("♻️ **ADMIN đã reset TOÀN BỘ!**")
    
    # Nếu không phải admin → Chỉ reset nhóm hiện tại
    else:
        reset_game_state(chat_id)
        
        # Chỉ reset stats của nhóm hiện tại
        str_chat_id = str(chat_id)
        if str_chat_id in stats:
            stats[str_chat_id] = {}
            save_stats(stats)
        
        await update.message.reply_text("✅ Trò chơi và bảng xếp hạng đã được reset **!")


def is_vietnamese(text):
    text = text.strip().lower()
    words = text.split()
    if len(words) != 2:
        return False
    if any(len(word) == 1 for word in words):
        return False
    if re.search(r'\d', text):
        return False
    vietnamese_pattern = r'^[a-zàáảãạâầấẩẫậăằắẳẵặèéẻẽẹêềếểễệìíỉĩịòóỏõọôồốổỗộơờớởỡợùúủũụưừứửữựỳýỷỹỵđ\s]+$'
    if not re.match(vietnamese_pattern, text):
        return False
    if words[0] not in TU_DON_LIST or words[1] not in TU_DON_LIST:
        return False
    return True


def contains_banned_words(text):
    words = text.lower().split()
    return any(word in GLOBAL_BANNED_WORDS for word in words)


def get_player_name(user, chat_id):
    if user.id in player_names[chat_id]:
        return player_names[chat_id][user.id]
    name = user.first_name
    if user.last_name:
        name += f" {user.last_name}"
    player_names[chat_id][user.id] = name
    return name


def get_player_username(user, chat_id):
    if user.username:
        player_usernames[chat_id][user.id] = user.username
        return user.username
    return "(chưa có username)"


async def start_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    # Nếu trò chơi đang diễn ra hoặc đang chờ người chơi nhập cụm đầu
    if in_game.get(chat_id) or waiting_for_phrase.get(chat_id):
        await update.message.reply_text(
            "⚠️ Trò chơi đang diễn ra, chưa kết thúc. Hãy ấn /luuy để hiểu thêm nhé!"
        )
        return

    reset_game_state(chat_id)  # Đặt lại toàn bộ trạng thái trò chơi
    game_start_time[chat_id] = datetime.now().strftime("%H:%M")

    await update.message.reply_text(
        "🎮 Trò chơi bắt đầu!\n"
        "👉 Gõ \u2003/join \u2003 Để tham gia.\n"
        "👉 Gõ \u2003/begin \u2003Khi đủ người, để bắt đầu.")


async def join_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user = update.effective_user
    if user.id in BANNED_USERS.get(chat_id, set()):
        await update.message.reply_text("🚫 Bạn đã bị cấm tham gia trò chơi.")
        return
    if user.id not in players.get(chat_id, []):
        players[chat_id].append(user.id)
        name = get_player_name(user, chat_id)
        username = get_player_username(user, chat_id)
        join_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        player_join_times[chat_id][user.id] = join_time
        save_player_to_excel(user.id, name, username, join_time)
        await update.message.reply_text(
            f"✅ {name} Đã tham gia! (Tổng: {len(players[chat_id])} Ng)")
    else:
        await update.message.reply_text("⚠️ Bạn đã tham gia rồi!")


async def begin_game(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id  # Lấy chat_id của nhóm
    global waiting_for_phrase, current_player_index, in_game

    # Kiểm tra nếu trò chơi đã bắt đầu
    if in_game.get(chat_id) or waiting_for_phrase.get(chat_id):
        await update.message.reply_text("⚠️ Trò chơi đã bắt đầu.")
        return

    if len(players.get(chat_id, [])) < 2:
        await update.message.reply_text(
            "❗ Cần ít nhất 2 người chơi để bắt đầu!")
        return

    in_game[chat_id] = True
    waiting_for_phrase[chat_id] = True
    current_player_index[chat_id] = 0
    user_id = players[chat_id][current_player_index[chat_id]]
    user = await context.bot.get_chat(user_id)

    await update.message.reply_text(
        f"✏️ {get_player_name(user, chat_id)}, Hãy nhập cụm từ đầu tiên:...\u2003\n"
        f"⏰ Bạn có 60 giây.")
    await start_turn_timer(context, chat_id)


async def play_word(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    global current_phrase, current_player_index, used_phrases, players, in_game, waiting_for_phrase, turn_timeout_task

    if not in_game.get(chat_id) or not waiting_for_phrase.get(
            chat_id) and current_phrase.get(chat_id) == "":
        return
    user = update.effective_user
    if user.id not in players.get(
            chat_id,
        []) or user.id != players[chat_id][current_player_index[chat_id]]:
        return

    text = update.message.text.strip().lower()
    if not is_vietnamese(text) or contains_banned_words(text):
        await eliminate_player(update, context, "Không hợp lệ!", chat_id)
        return

    if waiting_for_phrase.get(chat_id):
        current_phrase[chat_id] = text
        used_phrases[chat_id][text] = 1
        waiting_for_phrase[chat_id] = False
        await process_valid_word(update,
                                 context,
                                 text,
                                 is_first_word=True,
                                 chat_id=chat_id)
        return

    if text.split()[0] != current_phrase[chat_id].split()[-1]:
        await eliminate_player(
            update, context,
            f"Từ đầu phải là: '{current_phrase[chat_id].split()[-1]}'",
            chat_id)
        return

    if text in used_phrases[chat_id]:
        await eliminate_player(update, context, "Cụm từ đã dùng", chat_id)
        return

    used_phrases[chat_id][text] = 1
    current_phrase[chat_id] = text
    await process_valid_word(update, context, text, chat_id=chat_id)


async def process_valid_word(update,
                             context,
                             text,
                             is_first_word=False,
                             chat_id=None):
    global current_player_index, players, turn_timeout_task

    if turn_timeout_task.get(chat_id):
        turn_timeout_task[chat_id].cancel()
        turn_timeout_task[chat_id] = None

    if is_first_word:
        message = f"🎯 Từ bắt đầu: '{text}'\n\n"
    else:
        message = f"✅ {get_player_name(update.effective_user, chat_id)} Đã nối thành công!\n\n"

    current_player_index[chat_id] = (current_player_index[chat_id] + 1) % len(
        players[chat_id])

    if len(players[chat_id]) == 1:
        await announce_winner(update, context, chat_id)
        return

    current_word = current_phrase[chat_id].split()[-1]
    next_user = await context.bot.get_chat(
        players[chat_id][current_player_index[chat_id]])
    await update.message.reply_text(
        f"{message}"
        f"🔄 Lượt tiếp theo:\n"
        f"👉 Từ cần nối: 『\u2003{current_word}\u2003』\n"
        f"👤 Người chơi: {get_player_name(next_user, chat_id)}\n"
        f"⏳ Thời gian: 60 giây ")
    await start_turn_timer(context, chat_id)


async def eliminate_player(update, context, reason, chat_id):
    global players, current_player_index, turn_timeout_task

    user = update.effective_user
    name = get_player_name(user, chat_id)
    if user.id not in players.get(chat_id, []):
        return

    idx = players[chat_id].index(user.id)
    if turn_timeout_task.get(chat_id):
        turn_timeout_task[chat_id].cancel()
        turn_timeout_task[chat_id] = None

    await update.message.reply_text(f"❌ {name} Loại! Lý do: {reason}")
    players[chat_id].remove(user.id)

    if len(players[chat_id]) == 1:
        await announce_winner(update, context, chat_id)
        return

    if idx < current_player_index[chat_id]:
        current_player_index[chat_id] -= 1
    elif idx == current_player_index[chat_id] and current_player_index[
            chat_id] >= len(players[chat_id]):
        current_player_index[chat_id] = 0

    current_word = current_phrase[chat_id].split()[-1]
    next_user = await context.bot.get_chat(
        players[chat_id][current_player_index[chat_id]])
    await update.message.reply_text(
        f"👥 Người chơi còn lại: {len(players[chat_id])}\n"
        f"🔄 Lượt tiếp theo:\n"
        f"👉 Từ cần nối: 『\u2003{current_word}\u2003』\n"
        f"👤 Người chơi: {get_player_name(next_user, chat_id)}\n"
        f"⏳ Thời gian: 60 giây ")
    await start_turn_timer(context, chat_id)


async def announce_winner(update, context, chat_id):
    global stats, players

    if not players.get(chat_id):
        if update:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="🏁 Không có người chiến thắng.")
        reset_game_state(chat_id)
        return

    winner_id = players[chat_id][0]
    winner = await context.bot.get_chat(winner_id)
    name = get_player_name(winner, chat_id)

    # Lưu theo chat_id
    str_chat_id = str(chat_id)
    if str_chat_id not in stats:
        stats[str_chat_id] = {}
    stats[str_chat_id][name] = stats[str_chat_id].get(name, 0) + 1
    save_stats(stats)

    cid = update.effective_chat.id if update else chat_id
    await context.bot.send_message(
        chat_id=cid,
        text=f"🏆 CHIẾN THẮNG!🏆\n"
        f"👑 {name} -\u2003 Vô địch nối chữ!\n"
        f"📊 Số lần thắng:\u2003 {stats[str_chat_id][name]}")

    try:
        await context.bot.send_sticker(
            chat_id=cid,
            sticker=
            "CAACAgUAAxkBAAIBhWY9Bz7A0vjK0-BzFLEIF3qv7fBvAAK7AQACVp29V_R3rfJPL2MlNAQ"
        )
    except Exception as e:
        print(f"Lỗi gửi sticker thắng: {e}")

    reset_game_state(chat_id)


async def start_turn_timer(context, chat_id):
    global turn_timeout_task
    if turn_timeout_task.get(chat_id):
        turn_timeout_task[chat_id].cancel()
        turn_timeout_task[chat_id] = None

    turn_timeout_task[chat_id] = asyncio.create_task(
        turn_timer(context, chat_id))


async def turn_timer(context, chat_id):
    global players, current_player_index
    try:
        await asyncio.sleep(60)
        if not players.get(chat_id) or current_player_index[chat_id] >= len(
                players[chat_id]):
            return
        user_id = players[chat_id][current_player_index[chat_id]]
        user = await context.bot.get_chat(user_id)
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"⏰ {get_player_name(user, chat_id)} Hết giờ! Loại.")
        if user_id in players[chat_id]:
            players[chat_id].remove(user_id)
        if len(players[chat_id]) == 1:
            await announce_winner(None, context, chat_id)
            return
        if current_player_index[chat_id] >= len(players[chat_id]):
            current_player_index[chat_id] = 0
        current_word = current_phrase[chat_id].split()[-1]
        next_user = await context.bot.get_chat(
            players[chat_id][current_player_index[chat_id]])
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"🔄 Lượt tiếp theo:\n"
            f"👉 Từ cần nối: 『\u2003{current_word}\u2003』\n"
            f"👤 Người chơi: {get_player_name(next_user, chat_id)}\n"
            f"⏳ Thời gian: 60 giây")
        await start_turn_timer(context, chat_id)
    except asyncio.CancelledError:
        pass


async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat_id = str(update.effective_chat.id)

    if is_admin(user_id):
        if not stats:
            await update.message.reply_text(
                "📊 All chưa có ai thắng.")
            return

        message = "🏆 BẢNG XẾP HẠNG All 🏆\n\n"
        for group_id, group_stats in stats.items():
            message += f"📍 Nhóm {group_id}:\n"
            ranking = sorted(group_stats.items(),
                             key=lambda x: x[1],
                             reverse=True)
            for i, (name, wins) in enumerate(ranking[:10], 1):
                message += f"  {i}. {name}: {wins} Lần\n"
            message += "\n"
        await update.message.reply_text(message)
        return

    if chat_id not in stats or not stats[chat_id]:
        await update.message.reply_text("📊 Chưa có ai giành chiến thắng.")
        return

    ranking = sorted(stats[chat_id].items(), key=lambda x: x[1], reverse=True)
    message = "🏆 BẢNG XẾP HẠNG NHÓM 🏆\n\n"
    for i, (name, wins) in enumerate(ranking[:10], 1):
        message += f"{i}. {name}: {wins} Lần thắng\n"

    await update.message.reply_text(message)


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📖 HƯỚNG DẪN TRÒ CHƠI NỐI CHỮ\n\n"
        "🔹 /startgame - Bắt đầu trò chơi mới.\n"
        "🔹 /join - Tham gia trò chơi.\n"
        "🔹 /begin - Bắt đầu khi đủ người.\n"
        "🔹 /win - Xem bảng xếp hạng.\n"
        "🔹 /reset - Làm mới lại toàn bộ.\n"
        "🔹 /help - Xem hướng dẫn.\n\n"
        "📌 LUẬT CHƠI:\n"
        "- Mỗi cụm từ gồm 2 từ.\n"
        "- Nối từ cuối của cụm trước đó.\n"
        "- Không lặp lại cụm từ đã dùng.\n"
        "- Không dùng từ không phù hợp.\n"
        "- Mỗi lượt có 60 giây để trả lời.\n"
        "- Người cuối cùng còn lại sẽ chiến thắng.!\n"
        "👉 @xukaxuka2k1 code free,fastandsecure👈")


async def export_players(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    if not os.path.exists(EXCEL_FILE.format(chat_id=chat_id)):
        await update.message.reply_text("❌ Chưa có dữ liệu người chơi.")
        return

    with open(EXCEL_FILE.format(chat_id=chat_id), "rb") as f:
        await update.message.reply_document(document=f,
                                            filename=f"players_{chat_id}.xlsx")


async def clear_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    # Kiểm tra xem file Excel có tồn tại không
    if os.path.exists(EXCEL_FILE.format(chat_id=chat_id)):
        os.remove(EXCEL_FILE.format(chat_id=chat_id))
        await update.message.reply_text("🧹 File đã được xoá.")
    else:
        await update.message.reply_text("⚠️ Không tìm thấy file.")


async def add_word(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user

    if not is_admin(user.id):
        await update.message.reply_text("⚠️ Chưa thêm quyền.")
        return

    if not context.args:
        await update.message.reply_text("⚠️ Vui lòng nhập từ để thêm.")
        return

    new_word = context.args[0].strip().lower()

    if new_word in GLOBAL_BANNED_WORDS:
        await update.message.reply_text("⚠️ Từ này đã tồn tại.")
        return

    GLOBAL_BANNED_WORDS.add(new_word)
    save_banned_words()
    await update.message.reply_text(f"✅ Đã thêm từ cấm: '{new_word}' thành công.")


async def ban_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id  # Lấy chat_id của nhóm

    if not is_admin(update.effective_user.id):
        return await update.message.reply_text("❌ Chưa thêm quyền.")
    if not context.args:
        return await update.message.reply_text("⚠️ Cú pháp: /ban @username")

    username = context.args[0].lstrip('@')

    for uid, uname in player_usernames.get(chat_id, {}).items():
        if uname == username:
            BANNED_USERS.setdefault(chat_id, set()).add(uid)
            if uid in players.get(chat_id, []):
                players[chat_id].remove(uid)
            await update.message.reply_text(
                f"🚫 Đã ban {username} khỏi trò chơi.")
            return
    await update.message.reply_text("⚠️ Không tìm thấy người chơi đó.")


async def kick_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    if not is_admin(update.effective_user.id):
        return await update.message.reply_text(
            "❌ Bạn không có quyền sử dụng lệnh này.")
    if not context.args:
        return await update.message.reply_text("⚠️ Cú pháp: /kick @username")

    username = context.args[0].lstrip('@')

    for uid in players.get(chat_id, []):
        if player_usernames.get(chat_id, {}).get(uid) == username:
            players[chat_id].remove(uid)
            await update.message.reply_text(
                f"👢 Đã loại {username} khỏi trò chơi.")
            return
    await update.message.reply_text(
        "⚠️ Không tìm thấy người chơi đó trong danh sách.")


async def list_players(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    if not players.get(chat_id):
        return await update.message.reply_text("👥 Chưa có người chơi nào.")

    msg = "📋 Danh sách người chơi:\n"
    for i, uid in enumerate(players[chat_id], 1):
        name = player_names.get(chat_id, {}).get(uid, "(ẩn danh)")
        uname = player_usernames.get(chat_id, {}).get(uid, "")
        msg += f"{i}. {name} (@{uname})\n"

    await update.message.reply_text(msg)


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    # Kiểm tra quyền admin
    if not is_admin(update.effective_user.id):
        return await update.message.reply_text(
            "❌ Bạn không có quyền sử dụng lệnh này.")

    admin_commands = ("📜 **\u2003ADMIN\u2003** 📜\n"
                      "🔹 /ban @username - Cấm .\n"
                      "🔹 /kick @username - Kích.\n"
                      "🔹 /addword <từ> - Thêm từ:...\n"
                      "🔹 /reset - Làm mới lại toàn bộ." )

    await update.message.reply_text(admin_commands, parse_mode="Markdown")


async def luu_y(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    # Nội dung lưu ý
    note = ("⚠️ **Lưu ý** ⚠️\n\n"
            "🔹 **Trò chơi đang diễn ra chưa kết thúc.**\n"
            "🔹 **Khi nào kết thúc ấn [startgame] để tiếp tục chơi nhé!**")

    # Gửi tin nhắn lưu ý cho người chơi
    await update.message.reply_text(note)


async def unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id

    await update.message.reply_text(
        "❓ Lệnh không hợp lệ. Gõ /help để xem lệnh.\n\n"
        "🎮 game Caro:\u2003\u2003@Game_carobot\n"
        "🎮 Nối chữ:\u2003\u2003\u2003@noi_chu_bot"
    )

def main():
    TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("startgame", start_game))
    app.add_handler(CommandHandler("join", join_game))
    app.add_handler(CommandHandler("begin", begin_game))
    app.add_handler(CommandHandler("win", show_stats))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("fast", export_players))
    app.add_handler(CommandHandler("secure", clear_excel))
    app.add_handler(CommandHandler("addword", add_word))
    app.add_handler(CommandHandler("ban", ban_user))
    app.add_handler(CommandHandler("kick", kick_user))
    app.add_handler(CommandHandler("list", list_players))
    app.add_handler(CommandHandler("admin", admin_command))
    app.add_handler(CommandHandler("luuy", luu_y))
    app.add_handler(
        MessageHandler(filters.TEXT & (~filters.COMMAND), play_word))
    app.add_handler(MessageHandler(filters.COMMAND, unknown))

    print("Bot is running...")
    app.run_polling()
    
if __name__ == '__main__':
    main()
